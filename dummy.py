import time
start = time.time()
import bs4, requests
import lxml

extractedValueDict = []
movieNestedDict= {}
print("scraping web for 100 movies by IMDB start")
import urllib3
urlcount=0

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
urlIMDb=["https://www.imdb.com/search/title/?groups=top_250&sort=user_rating,desc&ref_=adv_prv",
	"https://www.imdb.com/search/title/?groups=top_250&sort=user_rating,desc&start=51&ref_=adv_nxt",
	"https://www.imdb.com/search/title/?groups=top_250&sort=user_rating,desc&start=101&ref_=adv_nxt",
	"https://www.imdb.com/search/title/?groups=top_250&sort=user_rating,desc&start=151&ref_=adv_nxt",
	"https://www.imdb.com/search/title/?groups=top_250&sort=user_rating,desc&start=201&ref_=adv_nxt"]
for url in urlIMDb:
    urlcount+=1
    tagName='h3'
    className="lister-item-header"
    # url = "https://medium.com/world-literature/creating-the-ultimate-list-100-books-to-read-before-you-die-45f1b722b2e5"
    res = requests.get(url, verify=False)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'lxml')
    # slno=soup.find_all("a",attrs={"class":"fi cn hx hy hz ia"})
    # bookAndAuthors = soup.find_all("strong", attrs={"class": "id ke"})
    ExtractedHTML = soup.find_all(tagName, attrs={"class": className})
    print(ExtractedHTML)
    #time.sleep(2)
    print("extracting data into extractedvaluedict")

    for i in range(0, len(ExtractedHTML)):
    
    #if len(ExtractedHTML[i].getText()):
        extractedValueDict.append(ExtractedHTML[i].getText())
        print(ExtractedHTML[i].getText())

extractedValueCount=0 
       #time.sleep(1)
for i in extractedValueDict:
        i="%r"%i
        print("i as raw text "+i)
        #time.sleep(2)
        
        movieNestedDict.update({str(i.split('\\n')[1].replace('.', '')): {'movie': i.split('\\n')[2], 'year': i.split('\\n')[3]}})
        
        print("extracted value count " + str(extractedValueCount))
        extractedValueCount+=1

print("opening new excel for book author rating")
import openpyxl
wb_objex = openpyxl.Workbook()
#wb_objex=openpyxl.load_workbook("top100MovieByIMDb.xlsx")
sheet_objex = wb_objex.active
sheet_objex.cell(row=1, column=1).value = 'Sl. No'
sheet_objex.cell(row=1, column=2).value = 'Movie Name'
sheet_objex.cell(row=1, column=3).value = 'Year of release'
sheet_objex.cell(row=1, column=4).value = 'Rating'

print("------writng list slno moive name year start------")

for bn, ba in movieNestedDict.items():
    print("book num", bn)

    col1 = bn  # booknum
    for key in ba:
         
        print(key + ":", ba[key])
        

        col2 = ba['movie']  # bookname
        col3 = ba['year']  # bookauthor
            # call write to text file function
        print("write to text for"+col1)
        #writeToTxtFile.writeToTxtFileFunction(col1, col2, col3)
        
        f = open("top100MovieByIMDb.txt", "a")
        f.write(str(col1) + '\t' + str(col2) + '\t' + col2 + '\n')
        f.close()
        
        sheet_objex.cell(row=int(col1)+1, column=1).value = str(col1)
        sheet_objex.cell(row=int(col1) + 1, column=2).value = str(col2)
        sheet_objex.cell(row=int(col1) + 1, column=3).value = str(col3)
        # sheet_objex.cell(row=col1+1,column=4).value='Rating'
print("------writing list slno book author to excel end------")
wb_objex.save("top100MovieByIMDb.xlsx")


print("urlcount",urlcount)
end = time.time()
print("time taken by program is:" + str(end - start))